from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def insertar_imagen (foto, celda):
    f_img = Image(foto)
    alto = f_img.height
    ancho = f_img.width
    factor = alto / 625
    f_img.width = ancho / factor
    f_img.height = alto /factor
    hoja.add_image(f_img, celda)

#RUTAS
ruta_sitios = "./archivos base/Sitios TP.xlsx"
ruta_ID = "./archivos base/pruebas.xlsx"

 #FOTOS
 #hoja1
img1 = "./sitio ejemplo/relevamiento/cota 0/fotos/1.jpeg"
img2 = "./sitio ejemplo/relevamiento/cota 0/fotos/2.jpeg"
img3 = "./sitio ejemplo/relevamiento/cota 0/fotos/3.jpeg"
img4 = "./sitio ejemplo/relevamiento/cota 0/fotos/4.jpeg"
img5 = "./sitio ejemplo/relevamiento/cota 0/fotos/5.jpeg"
img6 = "./sitio ejemplo/relevamiento/cota 0/fotos/6.jpeg"
 #hoja2
img7 = "./sitio ejemplo/relevamiento/cota 0/fotos/7.jpeg"
img8 = "./sitio ejemplo/relevamiento/cota 0/fotos/8.jpeg"
img9 = "./sitio ejemplo/relevamiento/cota 0/fotos/9.jpeg"
img10 = "./sitio ejemplo/relevamiento/cota 0/fotos/10.jpeg"
img11 = "./sitio ejemplo/relevamiento/cota 0/fotos/11.jpeg"
img12 = "./sitio ejemplo/relevamiento/cota 0/fotos/12.jpeg"

 #hoja3
img13 = "./fotos/13.jpeg"
img14 = "./fotos/14.jpeg"
img15 = "./fotos/15.jpeg"
img16 = "./fotos/16.jpeg"
img17 = "./fotos/17.jpeg"
img18 = "./fotos/18.jpeg"

 #fotos altura-1
img19 = "./fotos/altura/1.jpeg"
img20 = "./fotos/altura/2.jpeg"
img21 = "./fotos/altura/3.jpeg"
img22 = "./fotos/altura/4.jpeg"
img23 = "./fotos/altura/5.jpeg"
img24 = "./fotos/altura/6.jpeg"

 #fotos altura-2
img25 = "./fotos/altura/7.jpeg"
img26 = "./fotos/altura/8.jpeg"
img27 = "./fotos/altura/9.jpeg"
img28 = "./fotos/altura/10.jpeg"
img29 = "./fotos/altura/11.jpeg"
img30 = "./fotos/altura/12.jpeg"

#Lectura de archivos excel
wb_ID = load_workbook(ruta_ID)
ws_ID = wb_ID.active
wb_sitios = load_workbook(ruta_sitios)
ws_sitios = wb_sitios.active

#Listas de datos antenas
ant_list = ['Tribanda (Mod. RVVPX306.11R)' , 'Kathrein 80010764v01' , 'Twinbeam 2UNPX206.12R2','DualBeam HBXX-3817TB-VTM1','Parabola']
medida = ['1600x209x353','1400x299x152','1728x684x245',' 1390x301x181',' ф1200']
tecno = ['LTE 700 - AWS' , '2G 850-1900', '2G/3G 850-1900', '3G 1900', 'TX']
coax = ['6 / ф1/2' , '4 / ф1/2' , '8 / ф1/2' , '4 / ф1/2 ', '2 / ф7/8 ']
bat_list = ['PowerSafe', 'Narada' , 'Huawei']
bat_mod = ['12V170FS' , '12NDT190S' , 'ESM-48150B1']
EMG=' ' 
#busqueda de datos del sitio
ID_list = [celda[0].value for celda in ws_sitios['A1' : 'A34095']]
nom_list = [celda[0].value for celda in ws_sitios['B1' : 'B34095']]
EMG_list = [celda[0].value for celda in ws_sitios['D1' : 'D34095']]
lat_list =[celda[0].value for celda in ws_sitios['H1' : 'H34095']]
long_list = [celda[0].value for celda in ws_sitios['G1' : 'G34095']]
dir_list = [celda[0].value for celda in ws_sitios['S1' : 'S34095']]
prov_list = [celda[0].value for celda in ws_sitios['T1' : 'T34095']]
loc_list = [celda[0].value for celda in ws_sitios['U1' : 'U34095']]
az_list = [celda[0].value for celda in ws_sitios['K1' : 'K34095']]

ID = input ("ingrese el codigo del sitio: ")

#CARATULA
for i in range(0,34095):
    if ID == ID_list[i]:
        ws_ID['E3'] = ID_list[i]
        ws_ID['F3'] = nom_list[i]
        EMG = EMG_list[i]      
        ws_ID['E40'] = lat_list[i]
        ws_ID['I40'] = long_list[i]
        ws_ID['I30']= dir_list[i]
        ws_ID['E30'] = prov_list[i]
        ws_ID['E28'] = loc_list[i]
        AZA = az_list[i]
        i+=1
        AZB = az_list[i]
        i = i + 1 
        AZC = az_list[i]
        break

#RELEVAMIENTO COTA0-1
hoja = wb_ID [' RELEVAMIENTO COTA0-1']
wb_ID.active = hoja
insertar_imagen(img1, 'C14')
insertar_imagen(img2, 'N14')
insertar_imagen(img3, 'C51')
insertar_imagen(img4, 'N51')
insertar_imagen(img5, 'C88')
insertar_imagen(img6, 'N88')

#RELEVAMIENTO COTA0-2
hoja = wb_ID ['RELEVAMIENTO COTA0-2']
wb_ID.active = hoja
insertar_imagen(img7, 'C14')
insertar_imagen(img8, 'N14')
insertar_imagen(img9, 'C51')
insertar_imagen(img10, 'N51')
insertar_imagen(img11, 'C88')
insertar_imagen(img12, 'N88')    

#RELEVAMIENTO COTA0-3
hoja = wb_ID ['RELEVAMIENTO COTA0-3']
wb_ID.active = hoja
insertar_imagen(img13, 'C14')
insertar_imagen(img14, 'N14')
insertar_imagen(img15, 'C51')
insertar_imagen(img16, 'N51')
insertar_imagen(img17, 'C88')
insertar_imagen(img18, 'N88')  


#RELEVAMIENTO DE ESTRUCTURA
hoja = wb_ID ['RELEVAMIENTO DE ESTRUCTURA']
wb_ID.active = hoja
print ("****** RELEVAMIENTO DE ESTRUCTURA *******")
print (" ")
estructura = input ("tipo de estructura: ")
altura = input ("altura de estructura (en metros): ")
hoja['H13'] = estructura + " " + altura + "m"
print ("CARGAS DE ESTRUCTURA:     ")
i= int(24)
#rta = int (0)
while True:
    
    i = i + 1
    print ("ingrese el tipo de antena:")
    antena = int (input (" 0- Tribanda    1- DPDB   2- TwinBeam   3- DualBeam  4- Parabola "))
    h = input ("ingrese la altura de antana/parabola: ")
    
    if antena==0:
        #sector A
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZA
        hoja['F'+str(i)] = 'A'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        
        #sector B
        
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZB
        hoja['F'+str(i)] = 'B'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        #sector C
        
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZC
        hoja['F'+str(i)] = 'C'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
         
    if antena==1:
        #sector A
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZA
        hoja['F'+str(i)] = 'A'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        
        #sector B
     
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZB
        hoja['F'+str(i)] = 'B'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        #sector C
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZC
        hoja['F'+str(i)] = 'C'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        
    if antena==2:
        #sector A
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZA
        hoja['F'+str(i)] = 'A'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        #sector B
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZB
        hoja['F'+str(i)] = 'B'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        
        #sector C
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZC
        hoja['F'+str(i)] = 'C'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]

    if antena==3:
        #sector A
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZA
        hoja['F'+str(i)] = 'A'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        #sector B
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZB
        hoja['F'+str(i)] = 'B'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]
        i = i+1
        
        #sector C
        hoja['D'+str(i)] = h
        hoja['E'+str(i)] = AZC
        hoja['F'+str(i)] = 'C'
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['V'+str(i)] = 'TP'
        hoja['W'+str(i)] = tecno[antena]

    
    if antena==4:
        hoja['D'+str(i)] = h
        hoja['G'+str(i)] = ant_list[antena]
        hoja['M'+str(i)] = medida[antena]
        hoja['P'+str(i)] = coax[antena]
        hoja['W'+str(i)] = tecno[antena]
           
         
    rta = int (input (" ingresar otra carga?    1. SI   2.NO    "))
    if rta == int (2):
        break

insertar_imagen(img4, 'H55')

#RELEVAMIENTO DE ESTRUCTURA-01
hoja = wb_ID ['RELEVAMIENTO DE ESTRUCTURA-01']
wb_ID.active = hoja
insertar_imagen(img19, 'C14')
insertar_imagen(img20, 'N14')
insertar_imagen(img21, 'C51')
insertar_imagen(img22, 'N51')
insertar_imagen(img23, 'C88')
insertar_imagen(img24, 'N88')

#RELEVAMIENTO DE ESTRUCTURA-02
hoja = wb_ID ['RELEVAMIENTO DE ESTRUCTURA-02']
wb_ID.active = hoja
insertar_imagen(img25, 'C14')
insertar_imagen(img26, 'N14')
insertar_imagen(img27, 'C51')
insertar_imagen(img28, 'N51')
insertar_imagen(img29, 'C88')
insertar_imagen(img30, 'N88')


# MEMORIA DESCRIPTIVA
#hoja = wb_ID ['MEMORIA DESCRIPTIVA']
#wb_ID.active = hoja
#fxfc = input ('ingrese la cantidad de fxfc: ')
#fxcb = input ('ingrese la cantidad de fxcb: ')
#frig = input ('ingrese la cantidad de modulos frig/frij: ')
h_inst=input('altura de futura instalacion?:  ')
#frpa = 1
#fpfc = input('ingrese la cantida de modulos fpfc/h:  ')
eltek=input('tipo de eltek? 2 o 3  ')
rectif= input('cantidad de rectificadores: ')
bat_eltek = input('bancos de bat. dentro del eltek?: ')
print ('Marca de bat. dentro del eltek?:')
marca_bat_eltek = int (input('0- PowerSafe   1- Narada  2- Huawei'))
amp_eltek=input('amperaje?: ')

bat_18=input('cantidad de bancos en gabinete de baterias?: ')
print ('Marca de bat. dentro del Eltek 1.8m?:')
marca_bat_18 = int(input('0- PowerSafe   1- Narada  2- Huawei'))
amp_18=input('amperaje?: ')

#OVP = input('cantidad de OVP totales: ' )
#FSEP=input('cantidad de OVP FSEP: ') 
#hoja['C20']="un gabinete ELTEK-" + str(eltek) + " con " + str(rectif) + " Flatpack2 HE 48/2000 BtoF con " + str(bat_eltek) + " bancos PowerSafe de " + amp_eltek + "Ah"

# CONFIGURACION DE EQUIPOS
hoja = wb_ID ['CONFIGURACION DE EQUIPOS']
wb_ID.active = hoja
hoja['D32']=EMG+"B11"
hoja['D33']=EMG+"B12"
hoja['D34']=EMG+"B13"
hoja['D35']=EMG+"H11"
hoja['D36']=EMG+"H12"
hoja['D37']=EMG+"H13"
hoja['D38']=EMG+"N11"
hoja['D39']=EMG+"N12"
hoja['D40']=EMG+"N13"
hoja['D41']=EMG+"V11"
hoja['D42']=EMG+"V12"
hoja['D43']=EMG+"V13"
hoja['D44']=EMG+"V21"
hoja['D45']=EMG+"V22"
hoja['D46']=EMG+"V23"
hoja['D47']=EMG+"G11"
hoja['D48']=EMG+"G12"
hoja['D49']=EMG+"G13"
hoja['D50']=EMG+"L11"
hoja['D51']=EMG+"L12"
hoja['D52']=EMG+"L13"
hoja['D53']=EMG+"M11"
hoja['D54']=EMG+"M12"
hoja['D55']=EMG+"M13"



# CONFIGURACION DE RF
hoja = wb_ID ['CONFIGURACION DE RF']
wb_ID.active = hoja
#---Configuracion UMTS
hoja['D14'] = ID
hoja['E14'] = EMG
hoja['F14'] = EMG +'V11'
hoja['G14'] = estructura
hoja['H14'] = h_inst
hoja['M14'] = AZA
hoja['D15'] = ID
hoja['E15'] = EMG
hoja['F15'] = EMG +'V12'
hoja['G15'] = estructura
hoja['H15'] = h_inst
hoja['M15'] = AZB
hoja['D16'] = ID
hoja['E16'] = EMG
hoja['F16'] = EMG +'V13'
hoja['G16'] = estructura
hoja['H16'] = h_inst
hoja['M16'] = AZC
hoja['D17'] = ID
hoja['E17'] = EMG
hoja['F17'] = EMG +'V21'
hoja['G17'] = estructura
hoja['H17'] = h_inst
hoja['M17'] = AZA
hoja['D18'] = ID
hoja['E18'] = EMG
hoja['F18'] = EMG +'V22'
hoja['G18'] = estructura
hoja['H18'] = h_inst
hoja['M18'] = AZB
hoja['D19'] = ID
hoja['E19'] = EMG
hoja['F19'] = EMG +'V23'
hoja['G19'] = estructura
hoja['H19'] = h_inst
hoja['M19'] = AZC
#---Configuracion GSM
hoja['D25'] = ID
hoja['E25'] = EMG
hoja['F25'] = EMG +'H11'
hoja['G25'] = estructura
hoja['H25'] = h_inst
hoja['M25'] = AZA
hoja['D26'] = ID
hoja['E26'] = EMG
hoja['F26'] = EMG +'H12'
hoja['G26'] = estructura
hoja['H26'] = h_inst
hoja['M26'] = AZB
hoja['D27'] = ID
hoja['E27'] = EMG
hoja['F27'] = EMG +'H13'
hoja['G27'] = estructura
hoja['H27'] = h_inst
hoja['M27'] = AZC
hoja['D28'] = ID
hoja['E28'] = EMG
hoja['F28'] = EMG +'G11'
hoja['G28'] = estructura
hoja['H28'] = h_inst
hoja['M28'] = AZA
hoja['D29'] = ID
hoja['E29'] = EMG
hoja['F29'] = EMG +'G12'
hoja['G29'] = estructura
hoja['H29'] = h_inst
hoja['M29'] = AZB
hoja['D30'] = ID
hoja['E30'] = EMG
hoja['F30'] = EMG +'G13'
hoja['G30'] = estructura
hoja['H30'] = h_inst
hoja['M30'] = AZC
#---Configuracion LTE
hoja['D37'] = ID
hoja['E37'] = EMG
hoja['F37'] = EMG +'M11'
hoja['G37'] = estructura
hoja['H37'] = h_inst
hoja['M37'] = AZA
hoja['D38'] = ID
hoja['E38'] = EMG
hoja['F38'] = EMG +'M12'
hoja['G38'] = estructura
hoja['H38'] = h_inst
hoja['M38'] = AZB
hoja['D39'] = ID
hoja['E39'] = EMG
hoja['F39'] = EMG +'M13'
hoja['G39'] = estructura
hoja['H39'] = h_inst
hoja['M39'] = AZC
hoja['D40'] = ID
hoja['E40'] = EMG
hoja['F40'] = EMG +'L11'
hoja['G40'] = estructura
hoja['H40'] = h_inst
hoja['M40'] = AZA
hoja['D41'] = ID
hoja['E41'] = EMG
hoja['F41'] = EMG +'L12'
hoja['G41'] = estructura
hoja['H41'] = h_inst
hoja['M41'] = AZB
hoja['D42'] = ID
hoja['E42'] = EMG
hoja['F42'] = EMG +'L13'
hoja['G42'] = estructura
hoja['H42'] = h_inst
hoja['M42'] = AZC
hoja['D43'] = ID
hoja['E43'] = EMG
hoja['F43'] = EMG +'N11'
hoja['G43'] = estructura
hoja['H43'] = h_inst
hoja['M43'] = AZA
hoja['D44'] = ID
hoja['E44'] = EMG
hoja['F44'] = EMG +'N12'
hoja['G44'] = estructura
hoja['H44'] = h_inst
hoja['M44'] = AZB
hoja['D45'] = ID
hoja['E45'] = EMG
hoja['F45'] = EMG +'N13'
hoja['G45'] = estructura
hoja['H45'] = h_inst
hoja['M45'] = AZC
hoja['D46'] = ID
hoja['E46'] = EMG
hoja['F46'] = EMG +'B11'
hoja['G46'] = estructura
hoja['H46'] = h_inst
hoja['M46'] = AZA
hoja['D47'] = ID
hoja['E47'] = EMG
hoja['F47'] = EMG +'B12'
hoja['G47'] = estructura
hoja['H47'] = h_inst
hoja['M47'] = AZB
hoja['D48'] = ID
hoja['E48'] = EMG
hoja['F48'] = EMG +'B13'
hoja['G48'] = estructura
hoja['H48'] = h_inst
hoja['M48'] = AZC


# ANEXO DATOS DE ENTORNO
hoja = wb_ID ['ANEXO-DATOS DE ENTORNO']
wb_ID.active = hoja
hoja['D12'] = 'ELTEK'+ eltek

if marca_bat_eltek == 0:
    hoja['J12'] = bat_list[0]
    hoja['K12'] = bat_mod[0]
    hoja['L12'] = "BLOCK"
if marca_bat_eltek == 1:
    hoja['J12'] = bat_list[1]
    hoja['K12'] = bat_mod[1]
    hoja['L12'] = "BLOCK"
if marca_bat_eltek == 2:
    hoja['J12'] = bat_list[2]
    hoja['K12'] = bat_mod[2]
    hoja['L12'] = "MODULAR"
if marca_bat_18 == 0:
    hoja['J13'] = bat_list[0]
    hoja['K13'] = bat_mod[0]
    hoja['L13'] = "BLOCK"
if marca_bat_18 == 1:
    hoja['J13'] = bat_list[1]
    hoja['K13'] = bat_mod[1]
    hoja['L13'] = "BLOCK"
if marca_bat_18 == 2:
    hoja['J13'] = bat_list[2]
    hoja['K13'] = bat_mod[2]
    hoja['L13'] = "MODULAR"
    

hoja['M12'] = bat_eltek
hoja['N12'] = amp_eltek

hoja['M13'] = bat_18
hoja['N13'] = amp_18

insertar_imagen(img6, 'G29')

# ANEXO RESUMEN
hoja = wb_ID ['ANEXO-RESUMEN']
wb_ID.active = hoja
hoja['L14'] = h_inst
print (" ")
wb_ID.save("ingenieria.xlsx")
wb_sitios.save("sitios.xlsx")
print ("Ingenieria finalizada. Revisar!")


        
        
